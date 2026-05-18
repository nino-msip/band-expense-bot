import io
import os
import json
from collections import defaultdict
from datetime import datetime, timedelta

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build as build_service

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def _get_credentials():
    try:
        import streamlit as st
        creds_info = dict(st.secrets["gcp_service_account"])
    except Exception:
        creds_info = json.loads(os.environ["GOOGLE_CREDENTIALS_JSON"])
    return Credentials.from_service_account_info(creds_info, scopes=SCOPES)


def _get_client():
    return gspread.authorize(_get_credentials())


def _get_drive():
    return build_service("drive", "v3", credentials=_get_credentials())


def _create_folder(drive, folder_name: str) -> str:
    folder = drive.files().create(
        body={"name": folder_name, "mimeType": "application/vnd.google-apps.folder"},
        fields="id",
    ).execute()
    return folder["id"]


def _move_to_folder(drive, file_id: str, folder_id: str):
    file = drive.files().get(fileId=file_id, fields="parents").execute()
    previous_parents = ",".join(file.get("parents", []))
    drive.files().update(
        fileId=file_id,
        addParents=folder_id,
        removeParents=previous_parents,
        fields="id,parents",
    ).execute()


def _share_anyone(drive, file_id: str):
    drive.permissions().create(
        fileId=file_id,
        body={"type": "anyone", "role": "writer"},
    ).execute()


def _share_with_user(drive, file_id: str, email: str):
    drive.permissions().create(
        fileId=file_id,
        body={"type": "user", "role": "writer", "emailAddress": email},
        sendNotificationEmail=False,
    ).execute()


def upload_to_drive(xlsx_bytes: bytes, filename: str, folder_name: str) -> str:
    """XLSXをDriveフォルダにアップロードしてフォルダURLを返す。"""
    from googleapiclient.http import MediaIoBaseUpload

    drive = _get_drive()
    owner_email = _get_owner_email()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fn = folder_name.strip() or f"経費精算_{ts}"

    folder_id = _create_folder(drive, fn)

    media = MediaIoBaseUpload(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    file = drive.files().create(
        body={"name": filename, "parents": [folder_id]},
        media_body=media,
        fields="id",
    ).execute()

    # オーナーメールに直接共有 → 「共有アイテム」に表示される
    if owner_email:
        for fid in [folder_id, file["id"]]:
            _share_with_user(drive, fid, owner_email)
    else:
        _share_anyone(drive, folder_id)

    return f"https://drive.google.com/drive/folders/{folder_id}"


def _get_owner_email() -> str:
    email = ""
    try:
        import streamlit as st
        email = str(st.secrets.get("owner_email", ""))
    except Exception:
        pass
    if not email:
        email = os.environ.get("OWNER_EMAIL", "")
    return email


def cleanup_service_account_drive() -> tuple[int, int]:
    """サービスアカウントのDriveにある古いファイルをすべて削除する。(削除数, 合計数) を返す。"""
    drive = _get_drive()
    files = []
    page_token = None
    while True:
        kwargs = dict(
            q="trashed = false",
            fields="nextPageToken, files(id, name)",
            pageSize=100,
        )
        if page_token:
            kwargs["pageToken"] = page_token
        result = drive.files().list(**kwargs).execute()
        files.extend(result.get("files", []))
        page_token = result.get("nextPageToken")
        if not page_token:
            break

    deleted = 0
    for f in files:
        try:
            drive.files().delete(fileId=f["id"]).execute()
            deleted += 1
        except Exception:
            pass
    return deleted, len(files)


def _transfer_ownership(drive, file_id: str, owner_email: str):
    if not owner_email:
        return
    drive.permissions().create(
        fileId=file_id,
        body={"type": "user", "role": "owner", "emailAddress": owner_email},
        transferOwnership=True,
    ).execute()


def create_expense_report(name: str, address: str, items: list, folder_name: str = "") -> list:
    gc = _get_client()
    drive = _get_drive()
    owner_email = _get_owner_email()
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S")

    # インボイス番号ごとにグループ化（なければ "その他"）
    groups: dict[str, list] = defaultdict(list)
    for item in items:
        key = item.get("invoice_number") or "その他"
        groups[key].append(item)

    # フォルダ作成
    fn = folder_name.strip() or f"経費精算_{name}_{ts}"
    folder_id = _create_folder(drive, fn)
    _share_anyone(drive, folder_id)
    _transfer_ownership(drive, folder_id, owner_email)

    # 依頼ごとに新規スプレッドシートを作成してフォルダへ移動
    spreadsheet = gc.create(f"経費精算書_{name}_{ts}")
    _move_to_folder(drive, spreadsheet.id, folder_id)
    _transfer_ownership(drive, spreadsheet.id, owner_email)

    urls = []
    first = True
    for invoice_key, group_items in groups.items():
        if first:
            ws = spreadsheet.sheet1
            ws.update_title(invoice_key)
            ws.resize(rows=35, cols=10)
            first = False
        else:
            ws = spreadsheet.add_worksheet(title=invoice_key, rows=35, cols=10)
        total_row, note_row = _fill_form(ws, name, address, group_items, now)
        _format_sheet(ws, len(group_items), total_row, note_row)
        urls.append(
            f"https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit#gid={ws.id}"
        )

    return urls


def _fill_form(ws, name: str, address: str, items: list, now: datetime):
    issue_date = now.strftime("%Y年%-m月%-d日")

    updates = [
        {"range": "A2", "values": [["請求書（立替金清算書）"]]},
        {"range": "H4", "values": [[f"発行日　　{issue_date}"]]},
        {"range": "A6", "values": [["株式会社Vinyl Junkie Recordings"]]},
        {"range": "E6", "values": [["御中"]]},
        {"range": "H6", "values": [["氏名"]]},
        {"range": "I6", "values": [[name]]},
        {"range": "H7", "values": [["住所"]]},
        {"range": "I7", "values": [[address]]},
        {"range": "A10", "values": [["内容"]]},
        {"range": "E10", "values": [["税込金額"]]},
        {"range": "G10", "values": [["消費税　10％"]]},
        {"range": "I10", "values": [["備考"]]},
    ]

    total_amount = 0
    total_tax = 0
    invoice_notes = []

    for i, item in enumerate(items):
        row = 11 + i
        amount = item.get("amount_total") or 0
        tax = item.get("tax_amount") or round(amount * 10 / 110)
        date_str = item.get("date") or ""
        store = item.get("store_name") or ""
        desc = item.get("description") or ""
        invoice = item.get("invoice_number") or ""

        content = f"{date_str}　{desc}".strip("　 ") if date_str else desc

        biko = store  # 備考は店名のみ
        if store and invoice:
            invoice_notes.append(f"{store}（登録番号{invoice}）")

        updates += [
            {"range": f"A{row}", "values": [[content]]},
            {"range": f"E{row}", "values": [[amount]]},
            {"range": f"G{row}", "values": [[tax]]},
            {"range": f"I{row}", "values": [[biko]]},
        ]
        total_amount += amount
        total_tax += tax

    # 合計は16行目（枠内）、インボイス注記は19行目（欄外）固定
    TOTAL_ROW = 16
    NOTE_ROW = 19

    updates += [
        {"range": f"A{TOTAL_ROW}", "values": [["合　計"]]},
        {"range": f"E{TOTAL_ROW}", "values": [[total_amount]]},
        {"range": f"G{TOTAL_ROW}", "values": [[total_tax]]},
    ]

    invoice_notes = list(dict.fromkeys(invoice_notes))
    note_text = (
        "、".join(invoice_notes) + "への支払額として"
        if invoice_notes
        else "○○株式会社（登録番号T××××）への支払額として"
    )
    updates.append({"range": f"E{NOTE_ROW}", "values": [[note_text]]})

    ws.batch_update(updates, value_input_option="RAW")
    return TOTAL_ROW, NOTE_ROW


def _format_sheet(ws, num_items: int, total_row: int, note_row: int):
    """
    total_row: 合計行（枠内、通常16）
    note_row:  インボイス注記行（枠外 欄外、通常19）
    """
    sid = ws.id
    sp = ws.spreadsheet
    TSR = 9   # table start row (0-indexed) = row 10

    BLACK  = {"red": 0, "green": 0, "blue": 0}
    WHITE  = {"red": 1, "green": 1, "blue": 1}
    DARK   = {"red": 0.2, "green": 0.2, "blue": 0.2}
    GRAY   = {"red": 0.95, "green": 0.95, "blue": 0.95}
    BORDER = {"style": "SOLID_MEDIUM", "color": BLACK}
    THIN   = {"style": "SOLID", "color": {"red": 0.75, "green": 0.75, "blue": 0.75}}

    def rng(r1, c1, r2, c2):
        return {"sheetId": sid, "startRowIndex": r1, "endRowIndex": r2,
                "startColumnIndex": c1, "endColumnIndex": c2}

    total_idx = total_row - 1  # 0-indexed（合計行）
    note_idx  = note_row - 1   # 0-indexed（欄外注記行）

    requests = [
        # ── タイトル行 A2:J2 結合 ──
        {"mergeCells": {"range": rng(1, 0, 2, 10), "mergeType": "MERGE_ALL"}},

        # ── H4:J4 発行日エリア結合 ──
        {"mergeCells": {"range": rng(3, 7, 4, 10), "mergeType": "MERGE_ALL"}},

        # ── ヘッダ行 (row10) 各列結合 ──
        {"mergeCells": {"range": rng(TSR, 0, TSR+1, 4),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(TSR, 4, TSR+1, 6),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(TSR, 6, TSR+1, 8),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(TSR, 8, TSR+1, 10), "mergeType": "MERGE_ALL"}},

        # ── 合計行 各列結合（枠内最終行）──
        {"mergeCells": {"range": rng(total_idx, 0, total_idx+1, 4),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(total_idx, 4, total_idx+1, 6),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(total_idx, 6, total_idx+1, 8),  "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": rng(total_idx, 8, total_idx+1, 10), "mergeType": "MERGE_ALL"}},

        # ── インボイス注記行（欄外）E19:J19 結合 ──
        {"mergeCells": {"range": rng(note_idx, 4, note_idx+1, 10), "mergeType": "MERGE_ALL"}},

        # ── タイトル書式 ──
        {"repeatCell": {
            "range": rng(1, 0, 2, 10),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "CENTER",
                "verticalAlignment": "MIDDLE",
                "textFormat": {"bold": True, "fontSize": 16},
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)",
        }},

        # ── 宛先ボックス罫線（内部merge前に適用してセル単位で確実に設定）──
        {"updateBorders": {
            "range": rng(5, 0, 7, 7),
            "top": BORDER, "bottom": BORDER, "left": BORDER, "right": BORDER,
        }},

        # ── 氏名/住所ボックス罫線（内部merge前に適用）──
        {"updateBorders": {
            "range": rng(5, 7, 7, 10),
            "top": BORDER, "bottom": BORDER, "left": BORDER, "right": BORDER,
            "innerHorizontal": THIN, "innerVertical": THIN,
        }},

        # ── ヘッダセクション内部結合（罫線適用後）──
        # A6:D6 会社名
        {"mergeCells": {"range": rng(5, 0, 6, 4),  "mergeType": "MERGE_ALL"}},
        # I6:J6 氏名値
        {"mergeCells": {"range": rng(5, 8, 6, 10), "mergeType": "MERGE_ALL"}},
        # I7:J7 住所値
        {"mergeCells": {"range": rng(6, 8, 7, 10), "mergeType": "MERGE_ALL"}},

        # ── ヘッダセクション書式 ──
        {"repeatCell": {
            "range": rng(3, 7, 4, 10),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "RIGHT",
                "verticalAlignment": "MIDDLE",
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        }},
        {"repeatCell": {
            "range": rng(5, 0, 6, 4),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "LEFT",
                "verticalAlignment": "BOTTOM",
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        }},
        {"repeatCell": {
            "range": rng(5, 4, 6, 5),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "LEFT",
                "verticalAlignment": "BOTTOM",
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        }},
        {"repeatCell": {
            "range": rng(5, 7, 7, 8),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "LEFT",
                "verticalAlignment": "MIDDLE",
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        }},
        {"repeatCell": {
            "range": rng(5, 8, 7, 10),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "LEFT",
                "verticalAlignment": "MIDDLE",
            }},
            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        }},

        # ── テーブルヘッダ行書式（濃色背景・白文字）──
        {"repeatCell": {
            "range": rng(TSR, 0, TSR+1, 10),
            "cell": {"userEnteredFormat": {
                "backgroundColor": DARK,
                "horizontalAlignment": "CENTER",
                "verticalAlignment": "MIDDLE",
                "textFormat": {"bold": True, "foregroundColor": WHITE, "fontSize": 10},
            }},
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)",
        }},

        # ── 合計行書式（薄グレー背景・太字）──
        {"repeatCell": {
            "range": rng(total_idx, 0, total_idx+1, 10),
            "cell": {"userEnteredFormat": {
                "backgroundColor": GRAY,
                "textFormat": {"bold": True},
            }},
            "fields": "userEnteredFormat(backgroundColor,textFormat)",
        }},

        # ── テーブル全体の罫線（row10〜total_row、欄外の注記行は含まない）──
        {"updateBorders": {
            "range": rng(TSR, 0, total_idx+1, 10),
            "top": BORDER, "bottom": BORDER, "left": BORDER, "right": BORDER,
            "innerHorizontal": THIN, "innerVertical": THIN,
        }},

        # ── 金額列：右寄せ・通貨フォーマット（データ行 row11〜total_row-1）──
        {"repeatCell": {
            "range": rng(TSR+1, 4, total_idx, 8),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "RIGHT",
                "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
            }},
            "fields": "userEnteredFormat(horizontalAlignment,numberFormat)",
        }},
        # 合計行の金額も右寄せ
        {"repeatCell": {
            "range": rng(total_idx, 4, total_idx+1, 8),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "RIGHT",
                "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
            }},
            "fields": "userEnteredFormat(horizontalAlignment,numberFormat)",
        }},

        # ── 備考列（I-J）のデータ行：テキスト折り返し ──
        {"repeatCell": {
            "range": rng(TSR+1, 8, total_idx, 10),
            "cell": {"userEnteredFormat": {"wrapStrategy": "WRAP"}},
            "fields": "userEnteredFormat(wrapStrategy)",
        }},

        # ── インボイス注記（欄外 row19）：左寄せ・斜体 ──
        {"repeatCell": {
            "range": rng(note_idx, 4, note_idx+1, 10),
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "LEFT",
                "textFormat": {"italic": True, "fontSize": 9},
            }},
            "fields": "userEnteredFormat(horizontalAlignment,textFormat)",
        }},

        # ── タイトル行の高さ ──
        {"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "ROWS", "startIndex": 1, "endIndex": 2},
            "properties": {"pixelSize": 48},
            "fields": "pixelSize",
        }},

        # ── 列幅：10列均等（100px）──
        {"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 10},
            "properties": {"pixelSize": 100},
            "fields": "pixelSize",
        }},
    ]

    # ── データ行（row11〜total_row-1）の列結合 ──
    for i in range(total_row - 11):
        r = TSR + 1 + i
        requests += [
            {"mergeCells": {"range": rng(r, 0, r+1, 4),  "mergeType": "MERGE_ALL"}},
            {"mergeCells": {"range": rng(r, 4, r+1, 6),  "mergeType": "MERGE_ALL"}},
            {"mergeCells": {"range": rng(r, 6, r+1, 8),  "mergeType": "MERGE_ALL"}},
            {"mergeCells": {"range": rng(r, 8, r+1, 10), "mergeType": "MERGE_ALL"}},
        ]

    sp.batch_update({"requests": requests})


# ────────────────────────────────────────────────────────────
# XLSX生成（テンプレートベース）
# ────────────────────────────────────────────────────────────

def _unique_title(title: str, existing: list) -> str:
    if title not in existing:
        return title
    for n in range(2, 100):
        candidate = f"{title[:29]}({n})"
        if candidate not in existing:
            return candidate
    return title


def create_expense_xlsx(name: str, address: str, items: list) -> bytes:
    from openpyxl import Workbook as _WB
    wb = _WB()
    wb.remove(wb.active)

    now = datetime.now()
    issue_date = now.strftime("%Y年%-m月%-d日")
    created_titles: list[str] = []

    for i, item in enumerate(items):
        store = (item.get("store_name") or item.get("description") or f"明細{i+1}").strip()
        title = _unique_title(store[:31], created_titles)
        created_titles.append(title)
        ws = wb.create_sheet(title=title)
        _write_item_sheet(ws, name, address, item, issue_date)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_item_sheet(ws, name: str, address: str, item: dict, issue_date: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    # ── 列幅（Numbers原本に準拠）──
    widths = {"A": 7.7, "B": 8.5, "C": 8.5, "D": 8.5,
              "E": 10, "F": 10, "G": 10, "H": 8,
              "I": 8.9, "J": 17.7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── 行高 ──
    heights = {1: 18, 2: 24, 3: 24, 4: 18, 5: 18,
               6: 18, 7: 18, 8: 18, 9: 19,
               10: 18, 11: 19, 12: 18, 13: 18,
               14: 18, 15: 18, 16: 18, 17: 18, 18: 18}
    for r, h in heights.items():
        ws.row_dimensions[r].height = h

    # ── 罫線ヘルパー ──
    def thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def blk(left=False, right=False, top=False, bottom=False):
        med = Side(style="medium", color="000000")
        gray = Side(style="thin", color="BFBFBF")
        return Border(
            left=med if left else gray,
            right=med if right else gray,
            top=med if top else gray,
            bottom=med if bottom else gray,
        )

    # 全セルに薄い罫線（Numbers テーブルグリッド再現）
    for r in range(1, 19):
        for c in range(1, 11):
            ws.cell(r, c).border = thin()

    # ── タイトル ──
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "請求書（立替金清算書）"
    c.font = Font(size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── 発行日 ──
    ws["H4"].value = f"発行日　　{issue_date}"
    ws["H4"].alignment = Alignment(horizontal="left", vertical="center")

    # ── 宛先・氏名・住所 ──
    ws.merge_cells("A6:D6")
    ws["A6"].value = "株式会社Vinyl Junkie Recordings"
    ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
    ws["E6"].value = "御中"
    ws["E6"].alignment = Alignment(horizontal="left", vertical="center")

    ws["H6"].value = "氏名"
    ws["H6"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("I6:J6")
    ws["I6"].value = name
    ws["I6"].alignment = Alignment(horizontal="left", vertical="center")

    ws["H7"].value = "住所"
    ws["H7"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("I7:J8")
    ws["I7"].value = address
    ws["I7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── テーブルヘッダ行（row10）──
    ws.merge_cells("A10:D10")
    ws["A10"].value = "内容"
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E10:F10")
    ws["E10"].value = "税込金額"
    ws["E10"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G10:H10")
    ws["G10"].value = "消費税　10%"
    ws["G10"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("I10:J10")
    ws["I10"].value = "備考"
    ws["I10"].alignment = Alignment(horizontal="center", vertical="center")

    # ── データ行（row11）──
    amount = item.get("amount_total") or 0
    tax = item.get("tax_amount") or round(amount * 10 / 110)
    date_str = item.get("date") or ""
    desc = item.get("description") or ""
    store_name = item.get("store_name") or ""
    invoice = item.get("invoice_number") or ""
    content = f"{date_str}　{desc}".strip("　 ") if date_str else desc

    ws.merge_cells("A11:D11")
    ws["A11"].value = content
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("E11:F11")
    ws["E11"].value = amount
    ws["E11"].number_format = "#,##0"
    ws["E11"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("G11:H11")
    ws["G11"].value = tax
    ws["G11"].number_format = "#,##0"
    ws["G11"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("I11:J11")
    ws["I11"].value = store_name
    ws["I11"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # テーブル部分（rows 10-11）に黒の外枠 + 内側thin
    for r in range(10, 12):
        for c in range(1, 11):
            ws.cell(r, c).border = blk(
                left=(c == 1),
                right=(c == 10),
                top=(r == 10),
                bottom=(r == 11),
            )

    # ── インボイス注記（row13）──
    if invoice and store_name:
        note = f"{store_name}（登録番号{invoice}）への支払額として"
    elif invoice:
        note = f"（登録番号{invoice}）への支払額として"
    else:
        note = "○○株式会社（登録番号T××××）への支払額として"

    ws.merge_cells("E13:J13")
    ws["E13"].value = note
    ws["E13"].alignment = Alignment(horizontal="left", vertical="center")
